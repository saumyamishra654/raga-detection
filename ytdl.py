"""
Robust YouTube -> audio downloader driven from an Excel sheet.

Features:
- Reads /mnt/data/db.xlsx (change path/sheet_name at the top if needed).
- Accepts Excel hyperlink objects, HYPERLINK(...) formulas, and plain-text URLs.
- Names output files using the "Filename" column (sanitized), falling back to video title if absent.
- Converts audio to mp3 or wav via ffmpeg (yt-dlp postprocessor).
- Writes a CSV log with status (downloaded/skipped/failed) and error messages.
- Prints a short summary at end.
"""

import re
import os
import csv
import yt_dlp
from openpyxl import load_workbook
from html import unescape
import pathlib

# ==== SETTINGS ====
excel_path = '/Users/saumyamishra/Downloads/ragadb_tidy (4).xlsx'      # path to workbook (user uploaded)
sheet_name = "Database"        # sheet name inside workbook
link_column_header = "Link"           # header text containing the hyperlink
filename_column_header = "Filename"   # header text containing desired filenames
output_dir = "/Users/saumyamishra/Desktop/intern/summer25/RagaDetection/RagaDataset/fresh"   # where to save files
download_format = "mp3"               # "mp3" or "wav"
mp3_bitrate = "192"                   # kbps for mp3
log_csv = os.path.join(output_dir, "download_log.csv")
# ==================

os.makedirs(output_dir, exist_ok=True)


# helper: sanitize filename
def sanitize_filename(s, max_len=180):
    if s is None:
        return "untitled"
    # decode any HTML entities copied from Excel
    s = unescape(str(s))
    # replace slashes and control characters
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', s)
    s = s.strip()
    if len(s) > max_len:
        # keep extension area free, but here we just truncate
        s = s[:max_len]
    # collapse multiple underscores
    s = re.sub(r'__+', '_', s)
    if not s:
        return "untitled"
    return s

# helper: try to extract a URL from a text (handles HYPERLINK("url","text") formulas)
url_re = re.compile(r'https?://[^\s,;\'")\]]+')
def find_url_in_text(text):
    if not text:
        return None
    # if it's a HYPERLINK formula like =HYPERLINK("https://...","label")
    m = re.search(r'HYPERLINK\(\s*"([^"]+)"', str(text), flags=re.IGNORECASE)
    if m:
        return m.group(1)
    # plain text URL
    m2 = url_re.search(str(text))
    if m2:
        return m2.group(0)
    return None

# Load workbook
wb = load_workbook(excel_path, data_only=True)
if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
ws = wb[sheet_name]

# find header row and key column indexes
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
try:
    link_col_index = header_row.index(link_column_header) + 1
except ValueError:
    raise ValueError(f"Header '{link_column_header}' not found in first row headers: {header_row}")

try:
    filename_col_index = header_row.index(filename_column_header) + 1
except ValueError:
    # If no separate filename column exists, fall back to the link column.
    filename_col_index = link_col_index


# collect entries: (row_number, name_text, url)
entries = []
skipped_rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    link_cell = row[link_col_index - 1] if link_col_index - 1 < len(row) else None
    filename_cell = row[filename_col_index - 1] if filename_col_index - 1 < len(row) else None
    name_text = filename_cell.value if filename_cell is not None else None

    url = None
    # 1) check explicit Excel hyperlink object
    try:
        if link_cell and link_cell.hyperlink and getattr(link_cell.hyperlink, "target", None):
            url = link_cell.hyperlink.target
    except Exception:
        # some versions of openpyxl expose hyperlink differently; ignore errors
        url = None

    # 2) fallback: check the displayed cell value (covers HYPERLINK(...) formulas or raw URL text)
    if not url:
        url = find_url_in_text(link_cell.value if link_cell else None)

    # 3) fallback: scan other cells in same row for any URL (some spreadsheets put URL in adjacent column)
    if not url:
        for c in row:
            url = find_url_in_text(c.value)
            if url:
                break

    if url:
        entries.append((i, name_text, url))
    else:
        skipped_rows.append((i, name_text))

print(f"Found {len(entries)} candidate URLs; skipped {len(skipped_rows)} rows with no URL.")

# prepare log CSV
log_fieldnames = ["row", "name", "url", "status", "message", "saved_path"]
log_file = open(log_csv, "w", newline="", encoding="utf-8")
log_writer = csv.DictWriter(log_file, fieldnames=log_fieldnames)
log_writer.writeheader()

# main download loop
total = len(entries)
count_ok = 0
count_failed = 0
count_exist = 0

for idx, (rownum, name_text, url) in enumerate(entries, start=1):
    safe_name = sanitize_filename(name_text) if name_text else None

    # Check if file already exists to avoid re-processing
    if safe_name:
        existing_file = os.path.join(output_dir, f"{safe_name}.{download_format}")
        if os.path.exists(existing_file):
            print(f"[{idx}/{total}] Row {rownum}: '{name_text}' already exists. Skipping.")
            log_writer.writerow({
                "row": rownum,
                "name": name_text,
                "url": url,
                "status": "exists",
                "message": "File already exists",
                "saved_path": existing_file
            })
            count_exist += 1
            continue

    # use safe_name as base; if none provided yt-dlp will use video title later
    # create per-entry outtmpl so names don't clash
    if safe_name:
        outtmpl = os.path.join(output_dir, f"{safe_name}.%(ext)s")
    else:
        # fallback to yt-dlp title-based naming
        outtmpl = os.path.join(output_dir, "%(title)s.%(ext)s")

    ydl_opts = {
        "format": "bestaudio/best",
        "outtmpl": outtmpl,
        "noplaylist": True,
        "ignoreerrors": False,
        "quiet": False,
        "no_warnings": True,
        # do not overwrite existing file unless different extension
        "noprogress": False,
        "postprocessors": [
            {
                "key": "FFmpegExtractAudio",
                "preferredcodec": download_format,
                "preferredquality": mp3_bitrate if download_format == "mp3" else None
            }
        ],
        # keep temp files next to output
        "continuedl": True,
    }

    print(f"[{idx}/{total}] Row {rownum}: downloading '{name_text}' -> {url}")
    try:
        # instantiate per-entry YDL so outtmpl can be unique
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            # we try to see what filename would be produced to avoid re-downloading identical final file
            # but because yt-dlp may append info, we rely on filesystem check after download
            ydl.download([url])

        # attempt to find the saved file (could be .mp3 or .wav depending on format)
        expected_paths = []
        if safe_name:
            expected_paths.append(os.path.join(output_dir, f"{safe_name}.{download_format}"))
        # also check for any file in output_dir that startswith safe_name (some yt-dlp sanitization differs)
        if safe_name:
            pattern = f"{safe_name}"
            candidates = [p for p in os.listdir(output_dir) if p.startswith(pattern) and p.lower().endswith(download_format)]
            if candidates:
                saved = os.path.join(output_dir, candidates[0])
            else:
                # fallback: find most recent file in output directory with correct ext
                matches = [os.path.join(output_dir, p) for p in os.listdir(output_dir) if p.lower().endswith(download_format)]
                saved = max(matches, key=os.path.getmtime) if matches else ""
        else:
            matches = [os.path.join(output_dir, p) for p in os.listdir(output_dir) if p.lower().endswith(download_format)]
            saved = max(matches, key=os.path.getmtime) if matches else ""

        log_writer.writerow({
            "row": rownum,
            "name": name_text,
            "url": url,
            "status": "downloaded",
            "message": "",
            "saved_path": saved
        })
        log_file.flush()
        count_ok += 1

    except Exception as e:
        count_failed += 1
        err = str(e)
        print(f"ERROR downloading row {rownum}: {err}")
        log_writer.writerow({
            "row": rownum,
            "name": name_text,
            "url": url,
            "status": "failed",
            "message": err,
            "saved_path": ""
        })
        log_file.flush()

# write skipped rows to the log as well
for rownum, name_text in skipped_rows:
    log_writer.writerow({
        "row": rownum,
        "name": name_text,
        "url": "",
        "status": "skipped_no_url",
        "message": "",
        "saved_path": ""
    })

log_file.close()

print("Done.")
print(f"Total entries processed: {total}")
print(f"  Successfully downloaded: {count_ok}")
print(f"  Already existed: {count_exist}")
print(f"  Failed downloads: {count_failed}")
print(f"  Skipped (no URL found in row): {len(skipped_rows)}")
print(f"Log written to {log_csv}")
